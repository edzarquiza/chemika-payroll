import streamlit as st
import openpyxl
import re
import io
import pandas as pd
from datetime import datetime

st.set_page_config(page_title="Chemika Tools | Dexterous", page_icon="📊", layout="wide")


st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

html, body,
[data-testid="stAppViewContainer"],
[data-testid="stAppViewBlockContainer"],
.stApp {
    background-color: #ffffff !important;
    font-family: 'Inter', sans-serif;
}
[data-testid="block-container"] {
    background-color: #ffffff !important;
    padding-top: 0 !important;
    max-width: 1100px;
    margin: 0 auto;
}
p, span, li, td, th, caption,
[data-testid="stMarkdownContainer"] p,
[data-testid="stMarkdownContainer"] li,
[data-testid="stCaptionContainer"] p {
    color: #1a1a2e !important;
}
h1, h2, h3, h4 { color: #1a1a2e !important; font-weight: 700; }
[data-testid="stTabs"] [role="tab"] {
    color: #6b7280 !important;
    font-weight: 500;
    font-size: 0.95rem;
    border-bottom: 2px solid transparent !important;
}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] {
    color: #0d7e6e !important;
    border-bottom: 2px solid #0d7e6e !important;
    font-weight: 600;
}
[data-testid="stTabs"] [role="tab"]:hover { color: #0d7e6e !important; }
[data-testid="stTabs"] [data-baseweb="tab-highlight"] {
    background-color: #0d7e6e !important;
}
[data-testid="stTabs"] [data-baseweb="tab-border"] {
    background-color: #e5e7eb !important;
}
[data-testid="stMetric"] {
    background: #f0fdf9 !important;
    border: 1px solid #d1fae5 !important;
    border-radius: 10px !important;
    padding: 14px !important;
}
[data-testid="stMetric"] label {
    color: #374151 !important;
    font-size: 0.78rem !important;
    font-weight: 600 !important;
    text-transform: uppercase;
    letter-spacing: 0.04em;
}
[data-testid="stMetricValue"] { color: #0d7e6e !important; font-weight: 700 !important; }
[data-testid="stButton"] > button[kind="primary"],
.stDownloadButton > button,
[data-testid="stDownloadButton"] > button {
    background: linear-gradient(135deg, #0d9488, #1d4ed8) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    width: 100% !important;
}
[data-testid="stButton"] > button[kind="primary"]:hover,
[data-testid="stDownloadButton"] > button:hover { opacity: 0.88 !important; }
[data-testid="stFileUploaderDropzone"] {
    background: #f9fafb !important;
    border: 2px dashed #d1fae5 !important;
    border-radius: 10px !important;
}
[data-testid="stFileUploaderDropzone"] p,
[data-testid="stFileUploaderDropzone"] span,
[data-testid="stFileUploaderDropzone"] small { color: #374151 !important; }
[data-testid="stFileUploaderDropzone"] button {
    background: #f0fdf9 !important;
    border: 1px solid #0d9488 !important;
    color: #0d7e6e !important;
    border-radius: 6px !important;
}
.stTextInput > label, .stNumberInput > label {
    color: #374151 !important;
    font-weight: 500 !important;
    font-size: 0.875rem !important;
}
.stTextInput input, .stNumberInput input {
    background: #ffffff !important;
    color: #1a1a2e !important;
    border: 1px solid #d1d5db !important;
    border-radius: 6px !important;
}
.stTextInput input:focus, .stNumberInput input:focus {
    border-color: #0d9488 !important;
    box-shadow: 0 0 0 2px rgba(13,148,136,0.15) !important;
}
[data-testid="stExpander"] {
    border: 1px solid #e5e7eb !important;
    border-radius: 8px !important;
    background: #ffffff !important;
}
[data-testid="stExpander"] summary { color: #374151 !important; font-weight: 500 !important; }
.dex-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 18px 0 14px 0;
    border-bottom: 2px solid #e5e7eb;
    margin-bottom: 24px;
}
.dex-header-right { font-size: 0.8rem; color: #9ca3af; text-align: right; line-height: 1.6; }
.dex-badge {
    display: inline-block;
    background: linear-gradient(135deg, #0d9488, #1d4ed8);
    color: #ffffff;
    font-size: 0.7rem;
    font-weight: 600;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    padding: 3px 10px;
    border-radius: 20px;
    margin-top: 4px;
}
/* Reset button — secondary style */
[data-testid="stButton"] > button[kind="secondary"] {
    background: #ffffff !important;
    color: #6b7280 !important;
    border: 1.5px solid #d1d5db !important;
    border-radius: 8px !important;
    font-weight: 500 !important;
}
[data-testid="stButton"] > button[kind="secondary"]:hover {
    border-color: #e05555 !important;
    color: #e05555 !important;
    background: #fff5f5 !important;
}
</style>
""", unsafe_allow_html=True)

import base64 as _b64, os as _os

def _load_logo(path="logo.png"):
    if _os.path.exists(path):
        with open(path, "rb") as _f:
            return _b64.b64encode(_f.read()).decode()
    return ""

_logo_b64 = _load_logo("logo.png")
_logo_tag = (
    f'<img src="data:image/png;base64,{_logo_b64}" style="height:52px;object-fit:contain;" alt="Dexterous Group" />'
    if _logo_b64 else
    '<span style="font-size:1.1rem;font-weight:700;color:#1a1a2e;">Dexterous</span>'
)
st.markdown(
    f'''<div class="dex-header">
    {_logo_tag}
    <div class="dex-header-right">
        Chemika Tools<br>
        <span class="dex-badge">Internal &middot; Dexterous Group</span>
    </div>
</div>''',
    unsafe_allow_html=True,
)

tab1, tab2 = st.tabs(["📋  Payroll Timesheet Extractor", "🧾  Invoice TXT Formatter"])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — Payroll Timesheet Extractor
# ══════════════════════════════════════════════════════════════════════════════

EMPLOYEES = [
    {"row": 4,  "surname": "Ambrose",       "initial": "C", "category": "CASUAL",    "std_base": 165},
    {"row": 5,  "surname": "Bevan",          "initial": "H", "category": "PART-TIME", "std_base": 65},
    {"row": 6,  "surname": "Cantillon",      "initial": "J", "category": "",          "std_base": 162.5},
    {"row": 7,  "surname": "Charlton",       "initial": "H", "category": "",          "std_base": 162.5},
    {"row": 8,  "surname": "Cillian",        "initial": "A", "category": "",          "std_base": 162.5},
    {"row": 9,  "surname": "Darshika",       "initial": "",  "category": "SALARY",    "std_base": 162.5},
    {"row": 10, "surname": "Doherty",        "initial": "M", "category": "SALARY",    "std_base": 162.5},
    {"row": 11, "surname": "Doughty",        "initial": "S", "category": "CASUAL",    "std_base": 0},
    {"row": 12, "surname": "El Khoury",      "initial": "C", "category": "SALARY",    "std_base": 162.5},
    {"row": 13, "surname": "Emmanuel",       "initial": "M", "category": "SALARY",    "std_base": 162.5},
    {"row": 14, "surname": "Falzado",        "initial": "J", "category": "SALARY",    "std_base": 162.5},
    {"row": 15, "surname": "Horrigan",       "initial": "N", "category": "SALARY",    "std_base": 162.5},
    {"row": 16, "surname": "Jogia",          "initial": "V", "category": "SALARY",    "std_base": 162.5},
    {"row": 17, "surname": "Jones",          "initial": "A", "category": "SALARY",    "std_base": 162.5},
    {"row": 18, "surname": "Lauren",         "initial": "P", "category": "",          "std_base": 162.5},
    {"row": 19, "surname": "Lee",            "initial": "N", "category": "PART-TIME", "std_base": 108.3},
    {"row": 20, "surname": "LeStrange",      "initial": "S", "category": "SALARY",    "std_base": 162.5},
    {"row": 21, "surname": "LYAKHOVA",       "initial": "K", "category": "SALARY",    "std_base": 162.5},
    {"row": 22, "surname": "Manchanayake",   "initial": "T", "category": "PART-TIME", "std_base": 130},
    {"row": 23, "surname": "Mao",            "initial": "F", "category": "FULL-TIME", "std_base": 162.5},
    {"row": 24, "surname": "Micallef",       "initial": "C", "category": "SALARY",    "std_base": 162.5},
    {"row": 25, "surname": "Miller",         "initial": "R", "category": "SALARY",    "std_base": 162.5},
    {"row": 26, "surname": "Moeun",          "initial": "N", "category": "SALARY",    "std_base": 162.5},
    {"row": 27, "surname": "Parison",        "initial": "L", "category": "",          "std_base": 162.5},
    {"row": 28, "surname": "Piva",           "initial": "L", "category": "PART-TIME", "std_base": 130},
    {"row": 29, "surname": "Reardon",        "initial": "P", "category": "SALARY",    "std_base": 162.5},
    {"row": 30, "surname": "Rose",           "initial": "J", "category": "SALARY",    "std_base": 162.5},
    {"row": 31, "surname": "Sekuljica",      "initial": "D", "category": "FULL-TIME", "std_base": 162.5},
    {"row": 32, "surname": "Simpson",        "initial": "D", "category": "PART-TIME", "std_base": 97.5},
    {"row": 33, "surname": "Sor",            "initial": "S", "category": "CASUAL",    "std_base": 6},
    {"row": 34, "surname": "Watson",         "initial": "P", "category": "FULL-TIME", "std_base": 162.5},
]


def extract_numbers(s):
    if s is None:
        return 0
    result = "".join(c for c in str(s) if c.isdigit() or c == ".")
    return float(result) if result else 0


def safe_num(v):
    if v is None or str(v).strip().upper() == "N/A" or str(v).strip() == "":
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0


def extract_name_from_filename(filename):
    name_no_ext = filename.rsplit(".", 1)[0] if "." in filename else filename
    name_clean = name_no_ext.replace("_", " ").replace("-", "").strip()
    name_clean = re.sub(r"\s+", " ", name_clean)
    parts = name_clean.split()
    setstring = (parts[0] + " " + parts[1]) if len(parts) >= 2 else (parts[0] if parts else "")
    return setstring.strip()


def match_employee(setstring):
    setstring_clean = setstring.strip().lower()
    for emp in EMPLOYEES:
        surname = emp["surname"].strip()
        initial = emp["initial"].strip()
        if surname.lower() == setstring_clean:
            return emp
        if (surname + " " + initial).strip().lower() == setstring_clean:
            return emp
    return None


def parse_surname_initial(filename):
    name_no_ext = filename.rsplit(".", 1)[0] if "." in filename else filename
    name_clean = name_no_ext.replace("_", " ").replace("-", "").strip()
    name_clean = re.sub(r"\s+", " ", name_clean)
    parts = name_clean.split()
    month_names = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]
    name_parts = []
    for p in parts:
        if p.rstrip(".").lower() in month_names or p.isdigit():
            break
        name_parts.append(p)
    if len(name_parts) >= 2 and len(name_parts[-1]) == 1:
        return " ".join(name_parts[:-1]), name_parts[-1]
    elif name_parts:
        return " ".join(name_parts), ""
    return filename, ""


def process_file(file_bytes, filename):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    setstring = extract_name_from_filename(filename)
    emp = match_employee(setstring)
    is_new = emp is None
    warnings, errors = [], []

    sheet_name = ws.title.replace(",", "").strip()
    file_parts = setstring.lower().split()
    sheet_parts = sheet_name.lower().split()
    if file_parts and sheet_parts and file_parts[0] != sheet_parts[0]:
        errors.append(f"Filename says **{setstring}** but sheet is named **{ws.title}** — possible wrong file or copy error")

    g1 = ws.cell(1, 7).value
    header_std = None
    if g1:
        nums = re.findall(r"[\d.]+", str(g1))
        if nums:
            try:
                header_std = float(nums[0])
            except ValueError:
                pass

    AL = PL = 0.0
    prev_month_row = 49
    for row in range(6, 56):
        a = ws.cell(row, 1).value
        if a and "previous" in str(a).lower():
            prev_month_row = row
            break

    daily_ot_total = total_daily_hours = 0.0
    dates_seen, unsigned_working_days, day_mismatch_rows, weekend_no_ot = {}, [], [], []

    scan_end = min(prev_month_row, 49)
    for row in range(6, scan_end):
        a     = ws.cell(row, 1).value
        b     = ws.cell(row, 2).value
        c_val = ws.cell(row, 3).value
        d     = ws.cell(row, 4).value
        e     = ws.cell(row, 5).value
        f_ot  = ws.cell(row, 6).value
        c = safe_num(c_val)

        if c > 0:
            total_daily_hours += c
        if f_ot is not None and isinstance(f_ot, (int, float)) and f_ot > 0:
            daily_ot_total += f_ot

        if b and hasattr(b, "strftime"):
            date_str = b.strftime("%Y-%m-%d")
            if date_str in dates_seen:
                warnings.append(f"**Duplicate date** {date_str} at rows {dates_seen[date_str]} and {row}")
            else:
                dates_seen[date_str] = row

            if a and str(a).strip().lower() in ("monday","tuesday","wednesday","thursday","friday","saturday","sunday"):
                if str(a).strip().lower() != b.strftime("%A").lower():
                    day_mismatch_rows.append(row)

            actual_day_name = b.strftime("%A")
            if actual_day_name in ("Saturday","Sunday") and c > 0:
                if f_ot is None or (isinstance(f_ot, (int, float)) and f_ot == 0):
                    weekend_no_ot.append(f"Row {row}: {actual_day_name} {date_str} — {c}hrs worked, no OT in column F")

        if a and str(a).strip().lower() in ("monday","tuesday","wednesday","thursday","friday"):
            if c > 0 and (d is None or str(d).strip() == "") and (e is None or str(e).strip() == ""):
                unsigned_working_days.append(row)

        if d is None:
            continue
        d_str = str(d).strip()
        d_lower = d_str.lower()

        if d_str in ("AL", "A/L") or d_lower == "annual leave":
            AL += c
        elif (d_str == "SL" or d_lower in ("carers leave", "sick")) and setstring != "LeStrange S" and len(d_str) < 11:
            PL += c
        elif "leave" in d_lower and d_lower not in ("annual leave",):
            PL += c

        if setstring == "LeStrange S" and d_str == "SL" and len(d_str) < 11:
            cell = ws.cell(row, 4)
            if cell.fill and cell.fill.start_color:
                rgb = str(cell.fill.start_color.rgb) if cell.fill.start_color.rgb else ""
                if "FFFF00" in rgb or "ffff00" in rgb.lower():
                    PL += c

        if "sl" in d_lower:
            num = extract_numbers(d_str)
            if num > 0 and d_str != "SL":
                PL += num

    if day_mismatch_rows:
        sample = ", ".join(str(r) for r in day_mismatch_rows[:5])
        extra = "..." if len(day_mismatch_rows) > 5 else ""
        warnings.append(f"**Day/date mismatch** on {len(day_mismatch_rows)} row(s) — day names don't match calendar dates (rows {sample}{extra})")
    for item in weekend_no_ot[:3]:
        warnings.append(f"**Weekend work without OT**: {item}")
    if unsigned_working_days:
        sample = ", ".join(str(r) for r in unsigned_working_days[:5])
        extra = "..." if len(unsigned_working_days) > 5 else ""
        warnings.append(f"**Missing sign-in** on {len(unsigned_working_days)} weekday(s) with hours recorded (rows {sample}{extra})")

    std_hrs = ot10 = ot15 = ot20 = 0
    for row in range(2, ws.max_row + 1):
        a = ws.cell(row, 1).value
        if a is None:
            continue
        a_str = str(a).strip()
        if a_str == "STD MONTHLY HOURS":
            std_hrs = safe_num(ws.cell(row, 3).value)
        elif a_str == "O/T X 1.0":
            ot10 = safe_num(ws.cell(row, 3).value)
        elif a_str == "O/T X 1.5":
            ot15 = safe_num(ws.cell(row, 3).value)
        elif a_str == "O/T X 2.0":
            ot20 = safe_num(ws.cell(row, 3).value)

    if header_std is not None and std_hrs > 0 and abs(header_std - std_hrs) > 0.01:
        errors.append(f"**STD hours mismatch**: Header says **{header_std}** but summary row says **{std_hrs}**")
    if std_hrs == 0:
        errors.append("**STD MONTHLY HOURS is 0 or missing** — cannot verify hours")

    summary_ot = ot10 + ot15 + ot20
    if (daily_ot_total > 0 or summary_ot > 0) and abs(daily_ot_total - summary_ot) > 0.5:
        warnings.append(f"**OT mismatch**: Daily OT entries total **{daily_ot_total:.1f}hrs** but summary OT rows total **{summary_ot:.1f}hrs** (diff: {abs(daily_ot_total - summary_ot):.1f}hrs)")

    if std_hrs > 0 and total_daily_hours > 0 and total_daily_hours < std_hrs * 0.5:
        warnings.append(f"**Low hours**: Total daily hours (**{total_daily_hours:.1f}**) is less than half of STD (**{std_hrs}**) — timesheet may be incomplete")

    lsl_hrs = 0
    e64 = ws.cell(64, 5).value
    if e64 and "lsl accrual" in str(e64).strip().lower():
        lsl_hrs = safe_num(str(ws.cell(64, 6).value).replace("hrs", ""))
    else:
        for check_row in [67, 68, 66, 57]:
            f_val = ws.cell(check_row, 6).value
            if f_val is not None and str(f_val).strip() and str(f_val).strip().upper() != "N/A":
                lsl_hrs = safe_num(str(f_val).replace("hrs", ""))
                break

    compass = safe_num(ws.cell(67, 3).value)
    wb.close()

    if is_new:
        surname, initial = parse_surname_initial(filename)
        emp = {"row": None, "surname": surname, "initial": initial, "category": "", "std_base": 0, "is_new": True}

    return {
        "filename": filename, "name": setstring, "matched": True, "is_new": is_new,
        "employee": emp, "std_hrs": std_hrs, "lsl_hrs": lsl_hrs,
        "ot10": ot10, "ot15": ot15, "ot20": ot20,
        "al": AL, "pl": PL, "compass": compass,
        "warnings": warnings, "errors": errors,
    }


def build_output(results, month_label):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Main"

    ws.cell(1, 2, f"PAYROLL TIMESHEET - Month of {month_label}")
    ws.cell(1, 2).font = openpyxl.styles.Font(bold=True, size=14)

    headers = {
        1: "No", 2: "EMPLOYEE", 3: "", 4: "Start date", 5: "No. Yrs",
        6: "Category", 7: "Days p/w", 8: "STD HRS", 9: "LSL Hours",
        10: "Normal Hours", 11: "Car Allow", 12: "First Aid Allow",
        13: "O/T 1.0", 14: "O/T 1.5", 15: "O/T 2.0",
        16: "Travel @85c per km", 17: "Annual Leave", 18: "Personal/Sick Leave",
        19: "Compass Leave", 20: "LSL Leave", 21: "LWOP", 22: "Bonus"
    }
    hf = openpyxl.styles.Font(bold=True, size=10)
    hfill = openpyxl.styles.PatternFill("solid", fgColor="D9E1F2")
    for col, hdr in headers.items():
        cell = ws.cell(3, col, hdr)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", wrap_text=True)

    result_map, new_results = {}, []
    for r in results:
        if r["matched"]:
            if r.get("is_new"):
                new_results.append(r)
            else:
                result_map[r["employee"]["surname"].strip().lower()] = r

    for emp in EMPLOYEES:
        row = emp["row"]
        ws.cell(row, 1, row - 3)
        ws.cell(row, 2, emp["surname"])
        ws.cell(row, 3, emp["initial"])
        ws.cell(row, 6, emp["category"] if emp["category"] else None)
        key = emp["surname"].strip().lower()
        if key in result_map:
            r = result_map[key]
            ws.cell(row, 8,  r["std_hrs"])
            ws.cell(row, 9,  r["lsl_hrs"])
            ws.cell(row, 13, r["ot10"] if r["ot10"] != 0 else "N/A")
            ws.cell(row, 14, r["ot15"] if r["ot15"] != 0 else "N/A")
            ws.cell(row, 15, r["ot20"] if r["ot20"] != 0 else "N/A")
            ws.cell(row, 17, r["al"])
            ws.cell(row, 18, r["pl"])
            ws.cell(row, 19, r["compass"])

    last_row = max(e["row"] for e in EMPLOYEES)
    new_fill = openpyxl.styles.PatternFill("solid", fgColor="FFF2CC")
    for idx, r in enumerate(sorted(new_results, key=lambda x: x["employee"]["surname"])):
        row = last_row + 1 + idx
        emp_r = r["employee"]
        ws.cell(row, 1, row - 3)
        ws.cell(row, 2, emp_r["surname"])
        ws.cell(row, 3, emp_r["initial"])
        ws.cell(row, 8,  r["std_hrs"])
        ws.cell(row, 9,  r["lsl_hrs"])
        ws.cell(row, 13, r["ot10"] if r["ot10"] != 0 else "N/A")
        ws.cell(row, 14, r["ot15"] if r["ot15"] != 0 else "N/A")
        ws.cell(row, 15, r["ot20"] if r["ot20"] != 0 else "N/A")
        ws.cell(row, 17, r["al"])
        ws.cell(row, 18, r["pl"])
        ws.cell(row, 19, r["compass"])
        for col in range(1, 23):
            ws.cell(row, col).fill = new_fill

    total_rows = last_row + len(new_results)
    col_widths = {1:5,2:18,3:4,4:12,5:8,6:12,7:8,8:10,9:10,10:12,11:10,12:12,13:8,14:8,15:8,16:16,17:12,18:16,19:12,20:10,21:8,22:8}
    for col, w in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    for row in range(4, total_rows + 1):
        for col in [8, 9, 13, 14, 15, 17, 18, 19, 20]:
            cell = ws.cell(row, col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    ws_issues = wb.create_sheet("Data Issues")
    ih_font = openpyxl.styles.Font(bold=True, size=11, color="FFFFFF")
    ih_fill = openpyxl.styles.PatternFill("solid", fgColor="4472C4")
    for col_idx, hdr in enumerate(["Employee", "Severity", "Issue"], 1):
        cell = ws_issues.cell(1, col_idx, hdr)
        cell.font = ih_font
        cell.fill = ih_fill

    err_fill   = openpyxl.styles.PatternFill("solid", fgColor="FFC7CE")
    err_font   = openpyxl.styles.Font(color="9C0006")
    warn_fill  = openpyxl.styles.PatternFill("solid", fgColor="FFEB9C")
    warn_font  = openpyxl.styles.Font(color="9C6500")
    issue_row  = 2
    all_r = list(result_map.values()) + new_results
    for r in sorted(all_r, key=lambda x: x["employee"]["surname"]):
        emp_name = f"{r['employee']['surname']} {r['employee']['initial']}".strip()
        for err in r.get("errors", []):
            clean = re.sub(r"\*\*", "", err)
            ws_issues.cell(issue_row, 1, emp_name)
            ws_issues.cell(issue_row, 2, "ERROR")
            ws_issues.cell(issue_row, 3, clean)
            for c in range(1, 4):
                ws_issues.cell(issue_row, c).fill = err_fill
                ws_issues.cell(issue_row, c).font = err_font
            issue_row += 1
        for warn in r.get("warnings", []):
            clean = re.sub(r"\*\*", "", warn)
            ws_issues.cell(issue_row, 1, emp_name)
            ws_issues.cell(issue_row, 2, "WARNING")
            ws_issues.cell(issue_row, 3, clean)
            for c in range(1, 4):
                ws_issues.cell(issue_row, c).fill = warn_fill
                ws_issues.cell(issue_row, c).font = warn_font
            issue_row += 1

    if issue_row == 2:
        ws_issues.cell(2, 1, "No issues found")
        ws_issues.cell(2, 1).font = openpyxl.styles.Font(color="006100")

    ws_issues.column_dimensions["A"].width = 20
    ws_issues.column_dimensions["B"].width = 12
    ws_issues.column_dimensions["C"].width = 90

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


with tab1:
    # ── Session state for reset ───────────────────────────────────────────────
    if "payroll_reset" not in st.session_state:
        st.session_state.payroll_reset = 0

    st.markdown("Upload employee timesheet Excel files to generate the monthly payroll summary.")

    uploaded_files = st.file_uploader(
        "Upload employee timesheet files",
        type=["xlsx"],
        accept_multiple_files=True,
        help="Select all employee timesheet Excel files for the month",
        key=f"payroll_uploader_{st.session_state.payroll_reset}",
    )

    if uploaded_files:
        st.markdown(f"**{len(uploaded_files)} file(s) selected**")

        with st.expander("📁 Uploaded files", expanded=False):
            for f in sorted(uploaded_files, key=lambda x: x.name):
                name_key = extract_name_from_filename(f.name)
                emp = match_employee(name_key)
                if emp:
                    st.markdown(f"✅ `{f.name}` → **{emp['surname']} {emp['initial']}**")
                else:
                    surname, initial = parse_surname_initial(f.name)
                    st.markdown(f"🆕 `{f.name}` → **{surname} {initial}** (new employee — will be added automatically)")

        current_month = datetime.now().strftime("%B")
        month_label = st.text_input("Month label for report header", value=current_month)

        col_run, col_reset = st.columns([4, 1])
        with col_reset:
            if st.button("🔄 Reset", use_container_width=True, help="Clear all files and start over"):
                st.session_state.payroll_reset += 1
                st.rerun()
        with col_run:
            run_extraction = st.button("🚀 Extract & Generate Summary", type="primary", use_container_width=True)

        if run_extraction:
            results = []
            progress = st.progress(0, text="Processing files...")
            for i, f in enumerate(uploaded_files):
                progress.progress((i + 1) / len(uploaded_files), text=f"Processing {f.name}...")
                result = process_file(f.read(), f.name)
                results.append(result)
                f.seek(0)
            progress.empty()

            existing          = [r for r in results if r["matched"] and not r.get("is_new")]
            new_employees     = [r for r in results if r["matched"] and r.get("is_new")]
            total_errors      = sum(len(r.get("errors",   [])) for r in results)
            total_warnings    = sum(len(r.get("warnings", [])) for r in results)
            files_with_issues = sum(1 for r in results if r.get("errors") or r.get("warnings"))

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Files Processed",   len(results))
            col2.metric("Existing",          len(existing))
            col3.metric("New Employees",     len(new_employees))
            col4.metric("Files with Issues", files_with_issues)

            if new_employees:
                st.info(f"**{len(new_employees)} new employee(s)** detected — added to output (highlighted yellow):")
                for r in new_employees:
                    st.markdown(f"- 🆕 **{r['employee']['surname']} {r['employee']['initial']}** — STD HRS: {r['std_hrs']}")

            if total_errors > 0 or total_warnings > 0:
                st.markdown("### 🔍 Data Quality Report")
                st.markdown(f"Found **{total_errors} error(s)** and **{total_warnings} warning(s)** across {files_with_issues} file(s).")
                for r in sorted(results, key=lambda x: x["employee"]["surname"] if x["matched"] else x["name"]):
                    errs  = r.get("errors",   [])
                    warns = r.get("warnings", [])
                    if not errs and not warns:
                        continue
                    emp_label = f"{r['employee']['surname']} {r['employee']['initial']}".strip() if r["matched"] else r["name"]
                    badge = "🆕 " if r.get("is_new") else ""
                    with st.expander(f"{badge}{'🔴' if errs else '🟡'} {emp_label} — {len(errs)} error(s), {len(warns)} warning(s)", expanded=bool(errs)):
                        for err in errs:
                            st.error(f"🔴 {err}")
                        for warn in warns:
                            st.warning(f"🟡 {warn}")
            else:
                st.success("✅ No data quality issues detected across all files.")

            all_results = (sorted(existing, key=lambda x: x["employee"]["row"]) +
                           sorted(new_employees, key=lambda x: x["employee"]["surname"]))
            if all_results:
                st.markdown("### 📋 Extraction Preview")
                preview_data = []
                for r in all_results:
                    label = f"{r['employee']['surname']} {r['employee']['initial']}"
                    if r.get("is_new"):
                        label += " 🆕"
                    issues = len(r.get("errors", [])) + len(r.get("warnings", []))
                    preview_data.append({
                        "Employee": label,
                        "STD HRS": r["std_hrs"], "LSL Hrs": r["lsl_hrs"],
                        "O/T 1.0": r["ot10"],    "O/T 1.5": r["ot15"],   "O/T 2.0": r["ot20"],
                        "AL": r["al"],            "PL/SL": r["pl"],       "Compass": r["compass"],
                        "Issues": f"⚠️ {issues}" if issues > 0 else "✅",
                    })
                st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)

                output_xl       = build_output(results, month_label)
                output_filename = f"Payroll_Summary_{month_label}_{datetime.now().year}.xlsx"
                st.download_button(
                    label="📥 Download Payroll Summary",
                    data=output_xl,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )
                st.caption("The downloaded Excel includes a **Data Issues** sheet with all warnings and errors.")

    st.markdown("---")
    st.caption("Chemika Payroll Timesheet Extractor v2.0 · Data validation enabled")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — Invoice TXT Formatter
# ══════════════════════════════════════════════════════════════════════════════

def txt_format_date(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, pd.Timestamp):
        return val.strftime("%d/%m/%Y")
    try:
        return pd.to_datetime(val).strftime("%d/%m/%Y")
    except Exception:
        return str(val)


def txt_clean_num(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    f = float(val)
    return str(int(f)) if f == int(f) else str(f)


def build_txt(df: pd.DataFrame, memo: str, due_date: int, due_days: int, tax_code: str, account: str) -> bytes:
    TAB, CRLF = "\t", "\r\n"
    required = ["Date", "Sub Total", "GST", "Company Name", "Invoice Number"]
    for col in required:
        if col not in df.columns:
            raise ValueError(f"Missing required column: '{col}'")
    df = df.copy()
    if "Other" not in df.columns:
        df["Other"] = ""
    df = df.sort_values(
        by=["Company Name", "Invoice Number"],
        key=lambda col: col.astype(str) if col.name == "Company Name"
                        else pd.to_numeric(col, errors="coerce").fillna(0),
    )
    header = TAB.join(["Date","Sub Total","Other","GST","Company Name","Invoice Number",
                        "Memo","TT Ex GST","TT Inc GST","Due Date","Due Days","Tax Code","Account"])
    blank  = TAB * 12
    lines  = [header]
    for _, row in df.iterrows():
        date_str  = txt_format_date(row["Date"])
        sub_total = txt_clean_num(row["Sub Total"])
        other     = txt_clean_num(row.get("Other", ""))
        gst       = txt_clean_num(row["GST"])
        company   = str(row["Company Name"]).strip()
        invoice   = str(int(row["Invoice Number"])) if pd.notna(row["Invoice Number"]) else ""
        try:
            tt_ex_str  = txt_clean_num(float(row["Sub Total"]))
            tt_inc_str = txt_clean_num(float(row["Sub Total"]) + float(row["GST"]))
        except Exception:
            tt_ex_str, tt_inc_str = sub_total, ""
        lines.append(TAB.join([date_str, sub_total, other, gst, company, invoice,
                                memo, tt_ex_str, tt_inc_str,
                                str(due_date), str(due_days), tax_code, account]))
        lines.append(blank)
    return (CRLF.join(lines)).encode("utf-8")


with tab2:
    st.markdown("Upload an invoice spreadsheet and convert it to the tab-delimited TXT format for accounting import.")

    inv_file = st.file_uploader(
        "Upload invoice spreadsheet (.xlsx or .csv)",
        type=["xlsx", "csv"],
        key="invoice_uploader",
    )

    st.markdown("**Output settings**")
    c1, c2 = st.columns(2)
    with c1:
        inv_memo    = st.text_input("Memo",    value="Certificate of Analysis", key="inv_memo")
        inv_account = st.text_input("Account", value="4-1100",                  key="inv_account")
    with c2:
        inv_due_date = st.number_input("Due Date (days)", min_value=0, value=2,  step=1, key="inv_due_date")
        inv_due_days = st.number_input("Due Days",        min_value=0, value=30, step=1, key="inv_due_days")
        inv_tax_code = st.text_input("Tax Code", value="GST", key="inv_tax_code")

    st.markdown("---")

    if inv_file:
        try:
            df_inv = pd.read_csv(inv_file) if inv_file.name.endswith(".csv") else pd.read_excel(inv_file, sheet_name="Sheet1")
            df_inv.columns = [str(c).strip() for c in df_inv.columns]
            df_inv = df_inv.dropna(how="all")

            row_count = len(df_inv)
            companies = df_inv["Company Name"].nunique() if "Company Name" in df_inv.columns else 0
            inv_min   = int(df_inv["Invoice Number"].min()) if "Invoice Number" in df_inv.columns else "—"
            inv_max   = int(df_inv["Invoice Number"].max()) if "Invoice Number" in df_inv.columns else "—"

            ic1, ic2, ic3, ic4 = st.columns(4)
            ic1.metric("Invoices",      row_count)
            ic2.metric("Companies",     companies)
            ic3.metric("First Invoice", inv_min)
            ic4.metric("Last Invoice",  inv_max)

            with st.expander(f"Preview raw data ({row_count} rows)"):
                st.dataframe(df_inv, use_container_width=True, hide_index=True)

            txt_bytes    = build_txt(df_inv, inv_memo, inv_due_date, inv_due_days, inv_tax_code, inv_account)
            out_filename = inv_file.name.rsplit(".", 1)[0] + ".txt"

            with st.expander("Preview TXT output (first 8 rows)"):
                preview_lines = txt_bytes.decode("utf-8").replace("\r\n", "\n").split("\n")[:16]
                st.code("\n".join(preview_lines), language=None)

            st.download_button(
                label="📥 Download TXT File",
                data=txt_bytes,
                file_name=out_filename,
                mime="text/plain",
                type="primary",
                use_container_width=True,
                key="inv_download",
            )
            st.success(f"✅ **{row_count} invoices** processed · **{companies} companies** · Sorted by Company → Invoice Number · Download as `{out_filename}`")

        except Exception as e:
            st.error(f"**Error processing file:** {e}")
            st.exception(e)
    else:
        st.info("Upload an invoice spreadsheet above to begin. Must include columns: Date, Sub Total, GST, Company Name, Invoice Number.")

    st.markdown("---")
    st.caption("Chemika Invoice TXT Formatter · Replaces Chemika_Text_Format_Automation_v1_5.xlsm")
